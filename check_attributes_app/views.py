import datetime
import json
import uuid

from django.shortcuts import render
from .forms import UploadFileForm
from django.views import View
from django.http import HttpResponse, JsonResponse
import os
import openpyxl
from django.utils.encoding import escape_uri_path


# Create your views here.
def handle_uploaded_file(f):
    """
    Функция для загрузки файлов
    """
    with open(f"{f.name}", "wb+") as destination:
        for chunk in f.chunks():
            destination.write(chunk)


class IndexView(View):
    """Главная страница"""

    def get(self, request):
        form = UploadFileForm()  # Форма выгрузки файла
        content = {'form': form,
                   }
        resp = render(request, 'check_attributes_app/index.html', content)
        return resp

    def post(self, request):
        print(request.POST)
        # Удалять полностью пустые столбцы
        try:
            if request.POST['checkbox_columns']:
                checkbox_columns = True
            else:
                checkbox_columns = False
        except:
            checkbox_columns = False

        # Удалять полностью пустые строки
        try:
            if request.POST['checkbox_rows']:
                checkbox_rows = True
            else:
                checkbox_rows = False
        except:
            checkbox_rows = False

        # Закрашивать пустые ячейки
        try:
            if request.POST['checkbox_color']:
                checkbox_color = True
            else:
                checkbox_color = False
        except:
            checkbox_color = False

        # Не закрашивать полностью пустые столбцы
        try:
            if request.POST['checkbox_color_columns']:
                checkbox_color_columns = True
            else:
                checkbox_color_columns = False

        except:
            checkbox_color_columns = False

        print('checkbox_columns:', checkbox_columns)
        print('checkbox_rows:', checkbox_rows)
        print('checkbox_rows:', checkbox_color)
        print('checkbox_color_columns:', checkbox_color_columns)


        f = request.FILES["file"]
        extension = "." + str(f).split('.')[-1]
        file_name_input = str(f).split('.')[-2]
        path_to_input_file = os.path.join(f'{uuid.uuid4()}{extension}')

        print(path_to_input_file)
        with open(path_to_input_file, "wb") as destination:
            for chunk in f.chunks():
                destination.write(chunk)

        try:
            # Создаем новый файл
            new_excel_file = openpyxl.Workbook()
            ws_new_excel_file = new_excel_file.active
            ws_new_excel_file.title = 'Export'

            my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

            no_fill = openpyxl.styles.fills.PatternFill(fill_type=None)

            wb = openpyxl.load_workbook(filename=path_to_input_file, data_only=True)
            worksheet = wb.active
            book_len = worksheet.max_row
            book_width = worksheet.max_column

            # Удаляем столбцы
            input_columns = []
            count_of_null_columns = 0  # Счетчик пустых колонок
            for column in range(1, book_width + 1):
                count = 0
                for row in range(2, book_len + 1):
                    if not worksheet.cell(row, column).value:
                        count += 1
                print(f'Столбец {column} -пустых ячеек {count}')
                input_columns.append(str(worksheet.cell(1, column).value).replace('\n', ' '))
                if count + 1 == book_len:
                    count_of_null_columns += 1
                    print(f'Столбец {column} -все ячейки пустые')
                    if checkbox_columns:  # Удаляем столбцы, если чекбокс
                        worksheet.delete_cols(column)
            print('-----')
            print(f'Пустых колонок {count_of_null_columns}')

            # Удаляем строки
            count_of_null_rows = 0
            for row in range(2, book_len + 1):
                count = 0
                for column in range(1, book_width + 1):
                    if worksheet.cell(row, column).value:
                        count += 1
                # print(f'Строка {row} - Заполненных ячеек')
                if count == book_width:
                    count_of_null_rows += 1
                    print(f'Строка {row} - все ячейки заполнены')
                    if checkbox_rows:  # Удаляем строки, если чекбокс
                        worksheet.delete_rows(row)

            # Сохраняем во временный
            temp_file_path_full = 'temp.xlsx'
            wb.save(temp_file_path_full)

            wb = openpyxl.load_workbook(filename=temp_file_path_full, data_only=True)
            worksheet = wb.active
            book_len = worksheet.max_row
            book_width = worksheet.max_column

            # Красим
            count_fill = 0
            for column in range(1, book_width + 1):
                for row in range(2, book_len + 1):
                    if not worksheet.cell(row, column).value:
                        count_fill += 1
                        if checkbox_color:  # Закрашивание ячеек
                            worksheet.cell(row, column).fill = my_fill

            # Отбеливание ячеек
            if checkbox_color_columns:
                for column in range(1, book_width + 1):
                    count = 0
                    for row in range(2, book_len + 1):
                        if not worksheet.cell(row, column).value:
                            count += 1
                    if count + 1 == book_len:
                        for row in range(2, book_len + 1):
                            worksheet.cell(row, column).fill = no_fill

            file_to_export = f'export.xlsx'
            wb.save(file_to_export)
            print(request.POST)
            print(request.FILES)

            with open(file_to_export, 'rb') as fh:
                # Установка mimetype для правильной обработки браузером
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                response = HttpResponse(fh.read(), content_type=mime_type)
                response['Content-Disposition'] = 'attachment; filename=' + escape_uri_path(
                    f'{file_name_input}_{datetime.datetime.now().year}-{datetime.datetime.now().month}-{datetime.datetime.now().day}-{datetime.datetime.now().hour}:{datetime.datetime.now().minute}:{datetime.datetime.now().second}.xlsx')
            os.remove(path_to_input_file)
            os.remove(temp_file_path_full)

            return response
        except Exception as e:
            print(e)
            os.remove(path_to_input_file)
